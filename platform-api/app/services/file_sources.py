"""Helpers for uploaded-file data sources.

* ``compute_view_name`` mirrors the Teiid import servlet's naming convention so
  a filename maps to the same VDB view name used by the query builder.
* ``detect_column_types`` inspects a sample of an uploaded file and classifies
  each column as ``date``, ``currency``, ``number`` or ``string`` so the UI can
  format values (item 6) and so the type hints can later be pushed into the VDB.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Any

logger = logging.getLogger(__name__)

###############################################################################
# Upload cleaning - filename, column headers, and cell values
###############################################################################

# Characters that are reserved or problematic in SQL identifiers, Teiid view
# names, and file systems.  We replace them during upload so the user never has
# to deal with escaping issues.
_RESERVED_CHARS_RE = re.compile(r'[\$,#@!%\^&\*\(\)\+={}\[\]|\\:;"\'<>\?/~`]')
_MULTI_UNDERSCORES = re.compile(r"_+")
_LEADING_TRAILING_US = re.compile(r"^_|_$")

# SQL/Teiid reserved words that should not be used as identifiers.
_SQL_RESERVED: set[str] = {
    "select", "from", "where", "insert", "update", "delete", "drop", "create",
    "alter", "table", "index", "grant", "revoke", "and", "or", "not", "null",
    "true", "false", "order", "group", "by", "having", "limit", "offset",
    "join", "inner", "outer", "left", "right", "on", "as", "in", "between",
    "like", "is", "exists", "case", "when", "then", "else", "end", "union",
    "all", "distinct", "into", "values", "set", "default", "primary", "key",
    "foreign", "references", "constraint", "check", "unique", "cascade",
    "with", "recursive", "over", "partition", "row", "rows", "range",
    "window", "fetch", "first", "last", "next", "trigger",
    # Teiid-specific/date/type keywords that the Teiid parser treats as reserved.
    "year", "month", "day", "hour", "minute", "second", "quarter", "epoch",
    "date", "time", "timestamp", "timezone", "zone",
    "user", "role", "schema", "catalog", "domain",
    "convert", "cast", "extract", "trim", "leading", "trailing", "both",
    "substring", "position", "overlay", "escape", "matches", "similar",
    "string", "integer", "boolean", "double", "float", "decimal", "short",
    "long", "char", "varchar", "clob", "blob", "xml", "biginteger",
    "bigdecimal", "object", "variant", "json",
}


def sanitize_filename(filename: str) -> str:
    """Clean an uploaded filename: strip reserved chars, whitespace → _."""
    if "." in filename:
        base, ext = filename.rsplit(".", 1)
    else:
        base, ext = filename, ""
    base = _RESERVED_CHARS_RE.sub("_", base)
    base = base.replace(" ", "_")
    base = _MULTI_UNDERSCORES.sub("_", base)
    base = _LEADING_TRAILING_US.sub("", base)
    if not base:
        base = "file"
    if ext:
        ext = re.sub(r"[^a-zA-Z0-9]", "", ext)
        return f"{base}.{ext}"
    return base


def sanitize_column_name(name: str) -> str:
    """Clean a column header: strip reserved chars, trim whitespace → _."""
    clean = _RESERVED_CHARS_RE.sub("_", name.strip())
    clean = clean.replace(" ", "_")
    clean = _MULTI_UNDERSCORES.sub("_", clean)
    clean = _LEADING_TRAILING_US.sub("", clean)
    if not clean:
        clean = "col"
    if clean.lower() in _SQL_RESERVED:
        clean = f"{clean}_col"
    # Identifiers must not start with a digit
    if clean[0].isdigit():
        clean = f"col_{clean}"
    return clean


def sanitize_csv_content(content: bytes) -> bytes:
    """Clean CSV bytes in-place: sanitize column headers and strip currency
    symbols / whitespace from cell values so Teiid can auto-detect types.

    Returns the cleaned CSV as bytes (UTF-8).
    """
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample.splitlines()[0] else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return content

    # Clean headers
    header = [sanitize_column_name(h) for h in rows[0]]
    # Deduplicate
    seen: dict[str, int] = {}
    deduped: list[str] = []
    for h in header:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    header = deduped

    # Clean cell values: strip leading/trailing whitespace and currency symbols
    cleaned_rows: list[list[str]] = [header]
    for row in rows[1:]:
        cleaned: list[str] = []
        for cell in row:
            v = cell.strip()
            # Strip currency symbols but keep the numeric content
            v = re.sub(r"^[\$\u20ac\u00a3\u00a5]\s*", "", v)
            # Strip thousands-separator commas from numbers like "1,234.56"
            if re.match(r"^-?[\d,]+\.?\d*$", v):
                v = v.replace(",", "")
            cleaned.append(v)
        cleaned_rows.append(cleaned)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter)
    writer.writerows(cleaned_rows)
    return buf.getvalue().encode("utf-8")


def sanitize_xlsx_content(content: bytes) -> bytes:
    """Clean XLSX in-memory: sanitize column headers and strip currency
    symbols / whitespace from cell values.

    Returns cleaned CSV bytes (the xlsx→csv conversion is done here so the
    downstream file-import pipeline receives already-clean tabular data).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_header = next(rows_iter)
    except StopIteration:
        return content

    header = [sanitize_column_name(str(h) if h is not None else "col") for h in raw_header]
    # Deduplicate
    seen_h: dict[str, int] = {}
    deduped_h: list[str] = []
    for h in header:
        if h in seen_h:
            seen_h[h] += 1
            deduped_h.append(f"{h}_{seen_h[h]}")
        else:
            seen_h[h] = 0
            deduped_h.append(h)
    header = deduped_h

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for row in rows_iter:
        cleaned: list[str] = []
        for cell in row:
            if cell is None:
                cleaned.append("")
                continue
            v = str(cell).strip()
            v = re.sub(r"^[\$\u20ac\u00a3\u00a5]\s*", "", v)
            if re.match(r"^-?[\d,]+\.?\d*$", v):
                v = v.replace(",", "")
            cleaned.append(v)
        writer.writerow(cleaned)
    wb.close()
    return buf.getvalue().encode("utf-8")


def sanitize_excel_content(content: bytes, extension: str = "xlsx") -> bytes:
    """Clean an Excel workbook in-place and return sanitized .xlsx bytes.

    Headers are run through ``sanitize_column_name``, string cell values have
    currency symbols / thousands-separator commas stripped, and duplicate column
    names are deduplicated. The returned bytes are a valid Excel workbook the
    Teiid Excel translator can import, keeping the original .xlsx/.xlsm
    extension instead of flattening to CSV.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        header_cells = list(ws[1])
        seen_h: dict[str, int] = {}
        for cell in header_cells:
            if cell.value is None:
                continue
            raw = str(cell.value).strip()
            clean = sanitize_column_name(raw)
            if clean in seen_h:
                seen_h[clean] += 1
                clean = f"{clean}_{seen_h[clean]}"
            else:
                seen_h[clean] = 0
            cell.value = clean
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    v = cell.value.strip()
                    v = re.sub(r"^[\$\u20ac\u00a3\u00a5]\s*", "", v)
                    if re.match(r"^-?[\d,]+\.?\d*$", v):
                        v = v.replace(",", "")
                    cell.value = v
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _csv_bytes_to_xlsx(content: bytes, extension: str = "xlsx") -> bytes:
    """Convert CSV/TSV bytes to a sanitized Excel workbook."""
    import openpyxl

    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in text.splitlines()[0] else ","
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in csv.reader(io.StringIO(text), delimiter=delimiter):
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# Formats that the Teiid servlet cannot import natively and must be flattened to CSV.
_FLATTENED_FORMATS = frozenset({"json", "xml", "xls"})


def physical_file_name(file_name: str, source_format: str | None) -> str:
    """Return the on-disk filename for a source given its display filename.

    Excel (.xlsx, .xlsm), CSV, TSV and TXT keep their extension. JSON, XML and
    legacy .xls are normalized to .csv for the Teiid import servlet.
    """
    ext = (source_format or "").lower()
    if not ext or ext in _FLATTENED_FORMATS:
        if "." in file_name:
            return file_name.rsplit(".", 1)[0] + ".csv"
        return file_name + ".csv"
    return file_name


def candidate_physical_names(file_name: str, source_format: str | None) -> list[str]:
    """Possible on-disk filenames for a source, newest/primary first."""
    base = sanitize_filename(file_name)
    ext = (source_format or "").lower()
    candidates = []
    if ext in {"xlsx", "xlsm"}:
        candidates.append(base)
        # Legacy rows may have been flattened to .csv before Excel preservation.
        if "." in base:
            candidates.append(f"{base.rsplit('.', 1)[0]}.csv")
        else:
            candidates.append(f"{base}.csv")
    elif ext in _FLATTENED_FORMATS:
        if "." in base:
            candidates.append(f"{base.rsplit('.', 1)[0]}.csv")
        else:
            candidates.append(f"{base}.csv")
    else:
        candidates.append(base)
    return candidates


def prepare_upload_content(file_name: str, content: bytes) -> tuple[str, bytes, str | None]:
    """Sanitize an upload and return ``(physical_filename, content, source_format)``.

    ``physical_filename`` is the name actually written to disk and sent to the
    Teiid servlet; ``source_format`` is the original extension for display.
    """
    original_format = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else None
    clean_name = sanitize_filename(file_name)
    ext = clean_name.rsplit(".", 1)[-1].lower() if "." in clean_name else ""

    if ext == "xls":
        # Legacy binary Excel is flattened to CSV.
        content = sanitize_xlsx_content(content)
        clean_name = clean_name.rsplit(".", 1)[0] + ".csv"
    elif ext in {"xlsx", "xlsm"}:
        content = sanitize_excel_content(content, ext)
    elif ext in {"csv", "tsv", "txt"}:
        content = sanitize_csv_content(content)
    elif ext in {"json", "xml"}:
        clean_name, content = convert_to_csv_if_needed(clean_name, content)
    # Unknown extension is passed through.
    return clean_name, content, original_format


def prepare_replacement_content(
    incoming_file_name: str,
    content: bytes,
    target_physical_name: str,
) -> bytes:
    """Convert/sanitize ``content`` so it can be written as ``target_physical_name``."""
    target_ext = target_physical_name.rsplit(".", 1)[-1].lower() if "." in target_physical_name else ""
    incoming_ext = (
        sanitize_filename(incoming_file_name).rsplit(".", 1)[-1].lower()
        if "." in incoming_file_name
        else ""
    )

    if target_ext == incoming_ext:
        if target_ext in {"xlsx", "xlsm"}:
            return sanitize_excel_content(content, target_ext)
        # CSV/TSV/TXT replacements are staged/activated as-is; the Teiid
        # servlet sanitizes headers and values on import.
        return content

    if target_ext in {"xlsx", "xlsm"}:
        if incoming_ext in {"xlsx", "xlsm", "xls"}:
            return sanitize_excel_content(content, target_ext)
        # CSV/TSV/TXT/JSON/XML -> xlsx
        if incoming_ext in {"json", "xml"}:
            _, csv_bytes = convert_to_csv_if_needed(f"data.{incoming_ext}", content)
        elif incoming_ext in {"csv", "tsv", "txt"}:
            csv_bytes = sanitize_csv_content(content)
        else:
            csv_bytes = content
        return _csv_bytes_to_xlsx(csv_bytes, target_ext)

    if target_ext == "csv":
        if incoming_ext in {"xlsx", "xlsm", "xls"}:
            return sanitize_xlsx_content(content)
        if incoming_ext in {"json", "xml"}:
            _, csv_bytes = convert_to_csv_if_needed(f"data.{incoming_ext}", content)
            return csv_bytes
        if incoming_ext in {"csv", "tsv", "txt"}:
            return sanitize_csv_content(content)
        return content

    return content


_CURRENCY_NAME_HINTS = (
    "amount",
    "price",
    "cost",
    "revenue",
    "total",
    "balance",
    "salary",
    "fee",
    "payment",
    "charge",
    "subtotal",
    "tax",
    "usd",
    "dollar",
)
_CURRENCY_CHARS = re.compile(r"^[\$\u20ac\u00a3\u00a5]")
_DATE_NAME_HINTS = ("date", "_dt", "day", "month", "year", "timestamp", "time")


def compute_view_name(filename: str) -> str:
    """Return the VDB view name the servlet assigns to ``filename``.

    Applies the same sanitization used during upload so the view name is safe
    as a SQL identifier.
    """
    if "." in filename:
        base, ext = filename.rsplit(".", 1)
    else:
        base, ext = filename, ""
    base_name = _RESERVED_CHARS_RE.sub("_", base)
    base_name = base_name.replace(" ", "_")
    base_name = _MULTI_UNDERSCORES.sub("_", base_name)
    base_name = _LEADING_TRAILING_US.sub("", base_name)
    if not base_name:
        base_name = "file"
    extension = ext.upper() if ext else ""
    return f"{base_name}_{extension}" if extension else base_name


def _clean_number(value: str) -> str:
    """Strip currency symbols, thousands separators and whitespace."""
    v = value.strip()
    v = re.sub(r"[\$\u20ac\u00a3\u00a5,]", "", v)
    v = v.replace("(", "-").replace(")", "")
    return v.strip()


def _is_number(value: str) -> bool:
    v = _clean_number(value)
    if v in ("", "-", "."):
        return False
    try:
        float(v)
        return True
    except ValueError:
        return False


_DATE_PATTERNS = (
    re.compile(r"^\d{4}-\d{1,2}-\d{1,2}([ T]\d{1,2}:\d{2}(:\d{2})?)?$"),
    re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}([ T]\d{1,2}:\d{2}(:\d{2})?)?$"),
    re.compile(r"^\d{1,2}-\d{1,2}-\d{2,4}$"),
    re.compile(r"^\d{1,2}\.\d{1,2}\.\d{2,4}$"),
)


def _is_date(value: str) -> bool:
    v = value.strip()
    if not v or _is_number(v):
        return False
    return any(p.match(v) for p in _DATE_PATTERNS)


def _classify(name: str, values: list[str]) -> str:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "string"
    lname = name.lower()

    date_hits = sum(1 for v in non_empty if _is_date(str(v)))
    if date_hits / len(non_empty) >= 0.8:
        return "date"

    num_hits = sum(1 for v in non_empty if _is_number(str(v)))
    if num_hits / len(non_empty) >= 0.8:
        has_currency_char = any(_CURRENCY_CHARS.match(str(v).strip()) for v in non_empty)
        name_is_currency = any(h in lname for h in _CURRENCY_NAME_HINTS)
        if has_currency_char or name_is_currency:
            return "currency"
        return "number"

    # Name hint fall-backs when sampled data is sparse/ambiguous.
    if any(h in lname for h in _DATE_NAME_HINTS):
        return "date"
    return "string"


def _read_csv_sample(content: bytes, max_rows: int = 200) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample.splitlines()[0] else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return [], []
    header = [h.strip() for h in rows[0]]
    data = rows[1 : 1 + max_rows]
    return header, data


def _read_xlsx_sample(content: bytes, max_rows: int = 200) -> tuple[list[str], list[list[str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in header_row]
    data: list[list[str]] = []
    for i, r in enumerate(rows_iter):
        if i >= max_rows:
            break
        data.append([("" if c is None else str(c)) for c in r])
    wb.close()
    return header, data


def _flatten_value(value: Any) -> str:
    """Render a JSON/XML cell value as a flat string.

    Scalars become their string form; nested objects/arrays are serialized to
    compact JSON so the column still carries the data without breaking the
    tabular shape.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, str | int | float):
        return str(value)
    import json as _json

    return _json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _rows_to_csv(rows: list[dict[str, Any]]) -> bytes:
    """Serialize a list of record dicts to CSV bytes (union of keys as header).

    Column names are sanitized to be SQL-safe.
    """
    raw_columns: list[str] = []
    seen_raw: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen_raw:
                seen_raw.add(key)
                raw_columns.append(key)
    # Map raw keys to sanitized column names
    col_map: dict[str, str] = {}
    seen_clean: dict[str, int] = {}
    columns: list[str] = []
    for raw in raw_columns:
        clean = sanitize_column_name(raw)
        if clean in seen_clean:
            seen_clean[clean] += 1
            clean = f"{clean}_{seen_clean[clean]}"
        else:
            seen_clean[clean] = 0
        col_map[raw] = clean
        columns.append(clean)
    if not columns:
        columns = ["value"]
        raw_columns = ["value"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_flatten_value(row.get(raw, "")) for raw in raw_columns])
    return buf.getvalue().encode("utf-8")


def _json_to_rows(content: bytes) -> list[dict[str, Any]]:
    import json as _json

    data = _json.loads(content.decode("utf-8-sig"))
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        # Common API wrappers: use the first list-of-objects value if present,
        # otherwise treat the object itself as a single row.
        nested = next(
            (v for v in data.values() if isinstance(v, list) and v and all(
                isinstance(i, dict) for i in v
            )),
            None,
        )
        records = nested if nested is not None else [data]
    else:
        records = [{"value": data}]
    rows: list[dict[str, Any]] = []
    for rec in records:
        if isinstance(rec, dict):
            rows.append(rec)
        else:
            rows.append({"value": rec})
    return rows


def _xml_to_rows(content: bytes) -> list[dict[str, Any]]:
    import xml.etree.ElementTree as ET

    root = ET.fromstring(content)

    def _local(tag: str) -> str:
        return tag.rsplit("}", 1)[-1] if "}" in tag else tag

    def _element_to_row(el: ET.Element) -> dict[str, Any]:
        row: dict[str, Any] = {}
        for attr, val in el.attrib.items():
            row[_local(attr)] = val
        for child in el:
            key = _local(child.tag)
            if len(child) == 0:
                row[key] = (child.text or "").strip()
            else:
                row[key] = {
                    _local(gc.tag): (gc.text or "").strip() for gc in child
                }
        text = (el.text or "").strip()
        if text and not row:
            row["value"] = text
        return row

    children = list(root)
    # Find the dominant repeating child tag -> those become the rows.
    if children:
        from collections import Counter

        counts = Counter(_local(c.tag) for c in children)
        top_tag, top_n = counts.most_common(1)[0]
        if top_n > 1:
            return [_element_to_row(c) for c in children if _local(c.tag) == top_tag]
        # Single record: the root's children describe one row.
        return [_element_to_row(root)]
    return [_element_to_row(root)]


def display_source(
    physical_name: str, source_format: str | None
) -> tuple[str, str]:
    """Return ``(display_file_name, source_type)`` honoring the original upload.

    JSON/XML uploads are flattened to CSV, so the physical file on disk is
    ``foo.csv``. When ``source_format`` is recorded (e.g. ``"json"``) we present
    the original extension instead (``foo.json`` + type ``json``); otherwise we
    fall back to the on-disk extension.
    """
    ext = physical_name.rsplit(".", 1)[-1].lower() if "." in physical_name else ""
    if source_format:
        stem = (
            physical_name.rsplit(".", 1)[0]
            if "." in physical_name
            else physical_name
        )
        return f"{stem}.{source_format}", source_format
    return physical_name, (ext or "file")


def convert_to_csv_if_needed(filename: str, content: bytes) -> tuple[str, bytes]:
    """Convert JSON/XML uploads to CSV so they ride the existing file pipeline.

    Returns ``(filename, content)`` unchanged for non-JSON/XML inputs. For
    ``.json``/``.xml`` the content is flattened to CSV and the filename's
    extension is rewritten to ``.csv`` so the Teiid import servlet (which only
    parses CSV/TXT/Excel) treats it as a normal tabular data source. Raises
    ``ValueError`` with a friendly message if the file can't be parsed.
    """
    lower = filename.lower()
    if lower.endswith(".json"):
        try:
            rows = _json_to_rows(content)
        except Exception as exc:
            raise ValueError(f"Could not parse JSON file: {exc}") from exc
    elif lower.endswith(".xml"):
        try:
            rows = _xml_to_rows(content)
        except Exception as exc:
            raise ValueError(f"Could not parse XML file: {exc}") from exc
    else:
        return filename, content

    csv_bytes = _rows_to_csv(rows)
    base = filename.rsplit(".", 1)[0]
    return f"{base}.csv", csv_bytes


def detect_column_types(content: bytes, filename: str) -> list[dict[str, Any]]:
    """Detect a formatting type for each column of an uploaded file.

    Returns a list of ``{"name", "type"}`` where ``type`` is one of
    ``date | currency | number | string``.  Never raises — on any parse error
    it returns an empty list so upload is unaffected.
    """
    try:
        lower = filename.lower()
        if lower.endswith((".xlsx", ".xlsm", ".xls")):
            header, data = _read_xlsx_sample(content)
        else:
            header, data = _read_csv_sample(content)
    except Exception as exc:  # detection is best-effort
        logger.warning("Column type detection failed for %s: %s", filename, exc)
        return []

    if not header:
        return []

    result: list[dict[str, Any]] = []
    for idx, name in enumerate(header):
        col_values = [row[idx] for row in data if idx < len(row)]
        col_type = _classify(name, col_values)
        # The view normalizes header spaces to underscores; expose both so the
        # grid can match either the raw or normalized field name.
        result.append(
            {
                "name": name,
                "field": name.replace(" ", "_"),
                "type": col_type,
            }
        )
    return result
