
from __future__ import annotations

import csv
import io
import re

###############################################################################
# Upload cleaning - filename, column headers, and cell values
###############################################################################

# Characters that are reserved or problematic in SQL identifiers, Teiid view
# names, and file systems.  We replace them during upload so the user never has
# to deal with escaping issues.
_RESERVED_CHARS_RE = re.compile(r'[\$,#@!%\^&\*\(\)\+={}\[\]|\\:;"\'<>\?/~`]')
_MULTI_UNDERSCORES = re.compile(r"_+")
_LEADING_TRAILING_US = re.compile(r"^_|_$")

# SQL/Teiid reserved words that should not be used as identifiers.
_SQL_RESERVED: set[str] = {
    "select", "from", "where", "insert", "update", "delete", "drop", "create",
    "alter", "table", "index", "grant", "revoke", "and", "or", "not", "null",
    "true", "false", "order", "group", "by", "having", "limit", "offset",
    "join", "inner", "outer", "left", "right", "on", "as", "in", "between",
    "like", "is", "exists", "case", "when", "then", "else", "end", "union",
    "all", "distinct", "into", "values", "set", "default", "primary", "key",
    "foreign", "references", "constraint", "check", "unique", "cascade",
    "with", "recursive", "over", "partition", "row", "rows", "range",
    "window", "fetch", "first", "last", "next", "trigger",
    # Teiid-specific/date/type keywords that the Teiid parser treats as reserved.
    "year", "month", "day", "hour", "minute", "second", "quarter", "epoch",
    "date", "time", "timestamp", "timezone", "zone",
    "user", "role", "schema", "catalog", "domain",
    "convert", "cast", "extract", "trim", "leading", "trailing", "both",
    "substring", "position", "overlay", "escape", "matches", "similar",
    "string", "integer", "boolean", "double", "float", "decimal", "short",
    "long", "char", "varchar", "clob", "blob", "xml", "biginteger",
    "bigdecimal", "object", "variant", "json",
}


def sanitize_filename(filename: str) -> str:
    """Clean an uploaded filename: strip reserved chars, whitespace → _."""
    if "." in filename:
        base, ext = filename.rsplit(".", 1)
    else:
        base, ext = filename, ""
    base = _RESERVED_CHARS_RE.sub("_", base)
    base = base.replace(" ", "_")
    base = _MULTI_UNDERSCORES.sub("_", base)
    base = _LEADING_TRAILING_US.sub("", base)
    if not base:
        base = "file"
    if ext:
        ext = re.sub(r"[^a-zA-Z0-9]", "", ext)
        return f"{base}.{ext}"
    return base


def sanitize_column_name(name: str) -> str:
    """Clean a column header: strip reserved chars, trim whitespace → _."""
    clean = _RESERVED_CHARS_RE.sub("_", name.strip())
    clean = clean.replace(" ", "_")
    clean = _MULTI_UNDERSCORES.sub("_", clean)
    clean = _LEADING_TRAILING_US.sub("", clean)
    if not clean:
        clean = "col"
    if clean.lower() in _SQL_RESERVED:
        clean = f"{clean}_col"
    # Identifiers must not start with a digit
    if clean[0].isdigit():
        clean = f"col_{clean}"
    return clean


def sanitize_csv_content(content: bytes) -> bytes:
    """Clean CSV bytes in-place: sanitize column headers and strip currency
    symbols / whitespace from cell values so Teiid can auto-detect types.

    Returns the cleaned CSV as bytes (UTF-8).
    """
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample.splitlines()[0] else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return content

    # Clean headers
    header = [sanitize_column_name(h) for h in rows[0]]
    # Deduplicate
    seen: dict[str, int] = {}
    deduped: list[str] = []
    for h in header:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    header = deduped

    # Clean cell values: strip leading/trailing whitespace and currency symbols
    cleaned_rows: list[list[str]] = [header]
    for row in rows[1:]:
        cleaned: list[str] = []
        for cell in row:
            v = cell.strip()
            # Strip currency symbols but keep the numeric content
            v = re.sub(r"^[\$\u20ac\u00a3\u00a5]\s*", "", v)
            # Strip thousands-separator commas from numbers like "1,234.56"
            if re.match(r"^-?[\d,]+\.?\d*$", v):
                v = v.replace(",", "")
            cleaned.append(v)
        cleaned_rows.append(cleaned)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter)
    writer.writerows(cleaned_rows)
    return buf.getvalue().encode("utf-8")


def sanitize_xlsx_content(content: bytes) -> bytes:
    """Clean XLSX in-memory: sanitize column headers and strip currency
    symbols / whitespace from cell values.

    Returns cleaned CSV bytes (the xlsx→csv conversion is done here so the
    downstream file-import pipeline receives already-clean tabular data).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_header = next(rows_iter)
    except StopIteration:
        return content

    header = [sanitize_column_name(str(h) if h is not None else "col") for h in raw_header]
    # Deduplicate
    seen_h: dict[str, int] = {}
    deduped_h: list[str] = []
    for h in header:
        if h in seen_h:
            seen_h[h] += 1
            deduped_h.append(f"{h}_{seen_h[h]}")
        else:
            seen_h[h] = 0
            deduped_h.append(h)
    header = deduped_h

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for row in rows_iter:
        cleaned: list[str] = []
        for cell in row:
            if cell is None:
                cleaned.append("")
                continue
            v = str(cell).strip()
            v = re.sub(r"^[\$\u20ac\u00a3\u00a5]\s*", "", v)
            if re.match(r"^-?[\d,]+\.?\d*$", v):
                v = v.replace(",", "")
            cleaned.append(v)
        writer.writerow(cleaned)
    wb.close()
    return buf.getvalue().encode("utf-8")


def sanitize_excel_content(content: bytes, extension: str = "xlsx") -> bytes:
    """Clean an Excel workbook in-place and return sanitized .xlsx bytes.

    Headers are run through ``sanitize_column_name``, string cell values have
    currency symbols / thousands-separator commas stripped, and duplicate column
    names are deduplicated. The returned bytes are a valid Excel workbook the
    Teiid Excel translator can import, keeping the original .xlsx/.xlsm
    extension instead of flattening to CSV.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        header_cells = list(ws[1])
        seen_h: dict[str, int] = {}
        for cell in header_cells:
            if cell.value is None:
                continue
            raw = str(cell.value).strip()
            clean = sanitize_column_name(raw)
            if clean in seen_h:
                seen_h[clean] += 1
                clean = f"{clean}_{seen_h[clean]}"
            else:
                seen_h[clean] = 0
            cell.value = clean
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    v = cell.value.strip()
                    v = re.sub(r"^[\$\u20ac\u00a3\u00a5]\s*", "", v)
                    if re.match(r"^-?[\d,]+\.?\d*$", v):
                        v = v.replace(",", "")
                    cell.value = v
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
